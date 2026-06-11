"""
BRANDit Meeting Intelligence + Task Reminder System

Core flow:
1. Upload or record meeting audio/video
2. Convert audio with FFmpeg
3. Transcribe with Groq Whisper API
4. Optional speaker diarization with pyannoteAI hosted API
5. Extract minutes, decisions, action items, owners and timelines with Groq LLM
6. Store tasks in SQLite database
7. Calculate due dates and reminder dates automatically
8. Show Pending / Overdue / Due Today / Due Soon dashboard
9. Send reminder emails using Gmail App Password from Streamlit front-end

Run:
    streamlit run app.py --server.fileWatcherType none

Required .env or Streamlit secrets:
    GROQ_API_KEY=your_groq_key
    PYANNOTE_API_KEY=your_pyannote_key     # optional if speaker diarization is off
    GROQ_MODEL=llama-3.3-70b-versatile
    WHISPER_API_MODEL=whisper-large-v3-turbo

Important:
    For Gmail reminders, do NOT enter your normal Gmail password.
    Use a Google Gmail App Password. Enter it in the Streamlit front-end only.
"""

# ── Windows / OpenMP / Streamlit watcher fixes ───────────────────────────────
import os
os.environ.setdefault("KMP_DUPLICATE_LIB_OK", "TRUE")
os.environ.setdefault("STREAMLIT_SERVER_FILE_WATCHER_TYPE", "none")

import sys
if sys.platform == "win32":
    import asyncio
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

import warnings
warnings.filterwarnings("ignore", message=".*torch.classes.*")
warnings.filterwarnings("ignore", category=UserWarning)

# ── Imports ──────────────────────────────────────────────────────────────────
import html
import json
import re
import shutil
import sqlite3
import smtplib
import ssl
import subprocess
import tempfile
import time
import uuid
from datetime import datetime, timedelta, date
from email.message import EmailMessage
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests
import streamlit as st
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

DB_PATH = "brandit_tasks.db"

# ── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="BRANDit · Meeting Intelligence & Task Reminders",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Styling ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
:root{
    --bg:#050814;
    --panel:#0b1020;
    --panel-2:#11182a;
    --line:rgba(255,255,255,.10);
    --text:#e8edf7;
    --muted:#94a3b8;
    --teal:#20e3c2;
    --blue:#38bdf8;
    --navy:#07111f;
}
html,body,[class*="css"]{font-family:Inter,Segoe UI,Arial,sans-serif;}
.stApp{background:radial-gradient(circle at top right,rgba(32,227,194,.09),transparent 34%),linear-gradient(180deg,#060914 0%,#050814 100%);color:var(--text);}
[data-testid="stSidebar"]{background:linear-gradient(180deg,#07111f 0%,#090d1a 100%)!important;border-right:1px solid rgba(32,227,194,.13);}
.block-container{padding-top:1.3rem!important;max-width:1380px;}
.hero{position:relative;overflow:hidden;background:linear-gradient(135deg,rgba(10,20,35,.98),rgba(9,22,38,.94) 56%,rgba(5,38,44,.90));border:1px solid rgba(32,227,194,.16);border-radius:22px;padding:34px 38px;margin-bottom:24px;box-shadow:0 18px 45px rgba(0,0,0,.24);}
.hero:before{content:'';position:absolute;inset:0;background:linear-gradient(90deg,rgba(255,255,255,.04) 1px,transparent 1px),linear-gradient(180deg,rgba(255,255,255,.025) 1px,transparent 1px);background-size:84px 84px;opacity:.18;pointer-events:none;}
.hero>*{position:relative;z-index:1;}
.eyebrow{display:inline-flex;align-items:center;gap:8px;background:rgba(32,227,194,.10);border:1px solid rgba(32,227,194,.18);color:var(--teal);border-radius:7px;padding:7px 12px;font-size:.72rem;font-weight:850;letter-spacing:.14em;text-transform:uppercase;margin-bottom:18px;}
.hero h1{font-size:2.35rem;margin:0 0 12px;color:var(--teal);font-weight:900;letter-spacing:-.04em;}
.hero p{color:#d9e4ef;margin:0;font-size:1.02rem;line-height:1.65;max-width:920px;}
.card{background:rgba(255,255,255,.035);border:1px solid rgba(255,255,255,.09);border-radius:16px;padding:20px;margin-bottom:18px;box-shadow:0 12px 32px rgba(0,0,0,.16);}
.card-title{font-size:.76rem;font-weight:850;color:var(--teal);text-transform:uppercase;letter-spacing:.12em;margin-bottom:12px;}
.brand-card{background:linear-gradient(135deg,rgba(32,227,194,.16),rgba(56,189,248,.08));border:1px solid rgba(32,227,194,.22);border-radius:20px;padding:18px;margin:8px 0 20px;box-shadow:0 16px 38px rgba(0,0,0,.24);}
.brand-card h2{font-size:1.15rem;margin:0 0 8px;color:#ffffff;font-weight:850;letter-spacing:.01em;}
.brand-card p{font-size:.84rem;color:#cbd5e1;line-height:1.55;margin:0;}
.contact-strip{font-size:.75rem;color:#9fb7c8;border-top:1px solid rgba(255,255,255,.08);padding-top:12px;margin-top:12px;line-height:1.55;}
.workflow{display:flex;gap:10px;align-items:center;margin:0 0 22px;color:#64748b;font-size:.72rem;text-transform:uppercase;letter-spacing:.12em;flex-wrap:wrap;}
.workflow span{border:1px solid rgba(255,255,255,.08);background:rgba(255,255,255,.035);color:#a7b6c9;border-radius:9px;padding:8px 12px;text-transform:none;letter-spacing:0;font-size:.78rem;}
.workflow .active{background:rgba(32,227,194,.16);border-color:rgba(32,227,194,.32);color:var(--teal);font-weight:800;}
.feature-card{background:rgba(255,255,255,.035);border:1px solid rgba(255,255,255,.09);border-radius:16px;padding:18px;min-height:118px;}
.feature-card h3{font-size:1rem;margin:0 0 8px;color:#f8fafc;}
.feature-card p{font-size:.9rem;color:#aab8c9;line-height:1.55;margin:0;}
.metric-strip{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;margin:16px 0 24px;}
.metric-box{background:rgba(255,255,255,.035);border:1px solid rgba(255,255,255,.08);border-radius:14px;padding:16px;text-align:center;}
.metric-num{font-size:1.45rem;font-weight:850;color:var(--teal);line-height:1;}
.metric-label{font-size:.72rem;color:#8796ab;margin-top:6px;text-transform:uppercase;letter-spacing:.07em;}
.section-header{font-size:1.05rem;font-weight:850;color:#f1f5f9;margin:22px 0 12px;display:flex;align-items:center;gap:10px;}
.section-header:after{content:'';height:1px;flex:1;background:linear-gradient(90deg,rgba(32,227,194,.45),transparent);}
.transcript-box{background:#07101f;border:1px solid rgba(255,255,255,.09);border-radius:12px;padding:14px;max-height:430px;overflow-y:auto;}
.seg-row{display:grid;grid-template-columns:70px 110px 1fr;gap:10px;align-items:flex-start;padding:8px;border-bottom:1px solid rgba(255,255,255,.055);}
.seg-time{color:#7dd3fc;font-family:Consolas,monospace;font-size:.8rem;}
.seg-spk{color:var(--teal);font-weight:850;font-size:.82rem;white-space:nowrap;}
.seg-text{color:#dbe5ef;font-size:.9rem;line-height:1.55;}
.small-muted{color:#8ca0b8;font-size:.82rem;}
.status-ok{color:#34d399;font-weight:700;}
.status-bad{color:#fb7185;font-weight:700;}
.footer-bar{text-align:center;color:#7d90a8;font-size:.76rem;padding:14px 16px;margin:26px 0 8px;border:1px solid rgba(255,255,255,.08);background:rgba(255,255,255,.025);border-radius:14px;}
.footer-bar b{color:var(--teal);}
.stButton>button{background:linear-gradient(135deg,#20e3c2,#12b5cb)!important;color:#031018!important;border:0!important;border-radius:10px!important;font-weight:850!important;width:100%;}
.stDownloadButton>button{background:rgba(255,255,255,.04)!important;color:#d9f7f2!important;border:1px solid rgba(32,227,194,.18)!important;border-radius:10px!important;font-weight:750!important;width:100%;}
[data-testid="stFileUploader"]{background:rgba(255,255,255,.025)!important;border:2px dashed rgba(32,227,194,.28)!important;border-radius:14px!important;padding:8px;}
#MainMenu, footer{visibility:hidden;}
@media(max-width:900px){.metric-strip{grid-template-columns:repeat(2,1fr)}.seg-row{grid-template-columns:60px 95px 1fr}.hero h1{font-size:1.8rem}}
</style>
""", unsafe_allow_html=True)

# ── General Helpers ──────────────────────────────────────────────────────────
def get_secret(name: str, default: Optional[str] = None) -> Optional[str]:
    """Read from Streamlit secrets first, then .env / environment variables."""
    try:
        value = st.secrets.get(name)  # type: ignore[attr-defined]
        if value:
            return str(value)
    except Exception:
        pass
    return os.getenv(name, default)


def fmt_time(seconds: float) -> str:
    seconds = int(seconds or 0)
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    return f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"


def today_date() -> date:
    return datetime.now().date()


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def check_ffmpeg() -> bool:
    return shutil.which("ffmpeg") is not None


def safe_int(value, default: int = 5, min_value: int = 0, max_value: int = 365) -> int:
    try:
        result = int(value)
    except Exception:
        result = default
    return max(min_value, min(max_value, result))


def parse_assignee_email_mapping(mapping_text: str) -> Dict[str, str]:
    """
    Parse sidebar mapping like:
        Team A=teama@company.com
        Team B: teamb@company.com
        Speaker 01, speaker01@company.com
    """
    mapping: Dict[str, str] = {}
    for raw_line in (mapping_text or "").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        if "=" in line:
            name, email = line.split("=", 1)
        elif ":" in line:
            name, email = line.split(":", 1)
        elif "," in line:
            name, email = line.split(",", 1)
        else:
            continue

        name = name.strip()
        email = email.strip()
        if name and "@" in email:
            mapping[name.lower()] = email
    return mapping


def lookup_owner_email(owner: str, assignee_email_map: Optional[Dict[str, str]] = None) -> str:
    """Find an assignee email from owner/team/person name."""
    owner = (owner or "").strip()
    if not owner:
        return ""

    # If the LLM/transcript already gives an email in the owner field, use it.
    email_match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", owner)
    if email_match:
        return email_match.group(0)

    mapping = assignee_email_map or {}
    owner_key = owner.lower()
    if owner_key in mapping:
        return mapping[owner_key]

    # Fuzzy containment match: "Team A / Research" can match "Team A".
    for key, email in mapping.items():
        if key in owner_key or owner_key in key:
            return email

    return ""


def prepare_audio_for_apis(input_path: str, output_path: str) -> None:
    """Convert uploaded audio/video to 16 kHz mono FLAC."""
    if not check_ffmpeg():
        raise RuntimeError("FFmpeg not found. Add ffmpeg to packages.txt on Streamlit Cloud.")

    cmd = [
        "ffmpeg", "-y",
        "-i", input_path,
        "-vn",
        "-map", "0:a:0",
        "-ac", "1",
        "-ar", "16000",
        "-c:a", "flac",
        output_path,
    ]
    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if result.returncode != 0:
        raise RuntimeError("FFmpeg failed to prepare audio:\n" + result.stderr[-2000:])


# ── SQLite Task Database ─────────────────────────────────────────────────────
def init_db() -> None:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            meeting_id TEXT,
            meeting_title TEXT,
            source_file TEXT,
            task_description TEXT NOT NULL,
            owner TEXT,
            owner_email TEXT,
            priority TEXT,
            context TEXT,
            assigned_on TEXT,
            due_days INTEGER,
            due_date TEXT,
            reminder_date TEXT,
            status TEXT DEFAULT 'Pending',
            manager_email TEXT,
            created_at TEXT,
            last_reminder_sent_at TEXT,
            reminder_count INTEGER DEFAULT 0,
            last_assignee_reminder_sent_at TEXT,
            assignee_reminder_count INTEGER DEFAULT 0
        )
    """)
    # Migrate older local SQLite databases created by previous app versions.
    existing_columns = {row[1] for row in cur.execute("PRAGMA table_info(tasks)").fetchall()}
    migrations = {
        "owner_email": "ALTER TABLE tasks ADD COLUMN owner_email TEXT",
        "last_assignee_reminder_sent_at": "ALTER TABLE tasks ADD COLUMN last_assignee_reminder_sent_at TEXT",
        "assignee_reminder_count": "ALTER TABLE tasks ADD COLUMN assignee_reminder_count INTEGER DEFAULT 0",
    }
    for column_name, statement in migrations.items():
        if column_name not in existing_columns:
            cur.execute(statement)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS meeting_reports (
            meeting_id TEXT PRIMARY KEY,
            meeting_title TEXT,
            source_file TEXT,
            created_at TEXT,
            executive_summary TEXT,
            raw_analysis_json TEXT
        )
    """)
    conn.commit()
    conn.close()


def insert_meeting_report(meeting_id: str, analysis: dict, source_file: str) -> None:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        INSERT OR REPLACE INTO meeting_reports
        (meeting_id, meeting_title, source_file, created_at, executive_summary, raw_analysis_json)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        meeting_id,
        analysis.get("meeting_title", "Meeting"),
        source_file,
        datetime.now().isoformat(timespec="seconds"),
        analysis.get("executive_summary", ""),
        json.dumps(analysis, ensure_ascii=False),
    ))
    conn.commit()
    conn.close()


def save_tasks_from_analysis(
    analysis: dict,
    source_file: str,
    manager_email: str,
    reminder_before_days: int = 1,
    assignee_email_map: Optional[Dict[str, str]] = None,
) -> int:
    """
    Store LLM-extracted action items in SQLite.

    Example:
        Team A task due in 5 days
        assigned_on = today
        due_date = today + 5 days
        reminder_date = due_date - reminder_before_days
    """
    action_items = analysis.get("action_items", []) or []
    meeting_id = uuid.uuid4().hex[:12]
    meeting_title = analysis.get("meeting_title", "Meeting")
    assigned_on = today_date()

    insert_meeting_report(meeting_id, analysis, source_file)

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    inserted = 0

    for item in action_items:
        task = (item.get("action") or item.get("task") or "").strip()
        if not task:
            continue

        owner = item.get("owner", "TBD")
        owner_email = (item.get("owner_email") or item.get("assignee_email") or "").strip()
        if not owner_email:
            owner_email = lookup_owner_email(owner, assignee_email_map)

        due_days = safe_int(item.get("due_days", 5), default=5, min_value=0, max_value=365)
        due = assigned_on + timedelta(days=due_days)
        reminder = due - timedelta(days=reminder_before_days)
        if reminder < assigned_on:
            reminder = assigned_on

        cur.execute("""
            INSERT INTO tasks
            (meeting_id, meeting_title, source_file, task_description, owner, owner_email, priority,
             context, assigned_on, due_days, due_date, reminder_date, status,
             manager_email, created_at, last_reminder_sent_at, reminder_count,
             last_assignee_reminder_sent_at, assignee_reminder_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            meeting_id,
            meeting_title,
            source_file,
            task,
            owner,
            owner_email,
            item.get("priority", "Medium"),
            item.get("context", ""),
            assigned_on.isoformat(),
            due_days,
            due.isoformat(),
            reminder.isoformat(),
            "Pending",
            manager_email.strip(),
            datetime.now().isoformat(timespec="seconds"),
            None,
            0,
            None,
            0,
        ))
        inserted += 1

    conn.commit()
    conn.close()
    return inserted


def compute_task_bucket(row: dict) -> str:
    """Dynamic dashboard status based on current date."""
    if row.get("status") == "Completed":
        return "Completed"

    due = parse_date(row["due_date"])
    today = today_date()

    if due < today:
        return "Overdue"
    if due == today:
        return "Due Today"
    if due <= today + timedelta(days=3):
        return "Due Soon"
    return "Upcoming"


def fetch_tasks(status_filter: str = "All", manager_email: str = "") -> pd.DataFrame:
    init_db()
    conn = sqlite3.connect(DB_PATH)
    query = "SELECT * FROM tasks"
    params = []

    if manager_email.strip():
        query += " WHERE manager_email = ?"
        params.append(manager_email.strip())

    query += " ORDER BY due_date ASC, priority DESC, id DESC"

    df = pd.read_sql_query(query, conn, params=params)
    conn.close()

    if df.empty:
        return df

    df["Bucket"] = df.apply(lambda r: compute_task_bucket(r.to_dict()), axis=1)
    df["Days Left"] = df["due_date"].apply(lambda d: (parse_date(d) - today_date()).days)

    if status_filter != "All":
        df = df[df["Bucket"] == status_filter]

    return df


def update_task_status(task_id: int, status: str) -> None:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("UPDATE tasks SET status = ? WHERE id = ?", (status, task_id))
    conn.commit()
    conn.close()


def delete_task(task_id: int) -> None:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    conn.commit()
    conn.close()


def tasks_needing_reminders(manager_email: str = "") -> pd.DataFrame:
    """
    Select pending tasks whose reminder date has arrived.

    To avoid spamming:
    - A task receives max one reminder per day.
    - Reminder is sent if reminder_date <= today and status is Pending.
    """
    df = fetch_tasks(status_filter="All", manager_email=manager_email)
    if df.empty:
        return df

    today = today_date().isoformat()

    def should_remind(row) -> bool:
        if row["status"] == "Completed":
            return False
        if row["reminder_date"] > today:
            return False
        last = row.get("last_reminder_sent_at")
        if pd.isna(last) or not str(last).strip():
            return True
        try:
            return str(last)[:10] < today
        except Exception:
            return True

    return df[df.apply(should_remind, axis=1)].copy()


def mark_reminders_sent(task_ids: List[int]) -> None:
    if not task_ids:
        return

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    now = datetime.now().isoformat(timespec="seconds")
    cur.executemany(
        """
        UPDATE tasks
        SET last_reminder_sent_at = ?,
            reminder_count = COALESCE(reminder_count, 0) + 1
        WHERE id = ?
        """,
        [(now, int(task_id)) for task_id in task_ids],
    )
    conn.commit()
    conn.close()


def add_manual_task(
    meeting_title: str,
    task_description: str,
    owner: str,
    owner_email: str,
    priority: str,
    due_date_value: date,
    reminder_date_value: date,
    manager_email: str,
    context: str = "",
) -> None:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    assigned = today_date()
    due_days = max(0, (due_date_value - assigned).days)
    cur.execute("""
        INSERT INTO tasks
        (meeting_id, meeting_title, source_file, task_description, owner, owner_email, priority,
         context, assigned_on, due_days, due_date, reminder_date, status,
         manager_email, created_at, last_reminder_sent_at, reminder_count,
         last_assignee_reminder_sent_at, assignee_reminder_count)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        "manual-" + uuid.uuid4().hex[:8],
        meeting_title or "Manual Task",
        "Manual Entry",
        task_description,
        owner or "TBD",
        owner_email.strip(),
        priority,
        context,
        assigned.isoformat(),
        due_days,
        due_date_value.isoformat(),
        reminder_date_value.isoformat(),
        "Pending",
        manager_email.strip(),
        datetime.now().isoformat(timespec="seconds"),
        None,
        0,
        None,
        0,
    ))
    conn.commit()
    conn.close()


# ── Notification Helpers ─────────────────────────────────────────────────────
def build_reminder_body(tasks_df: pd.DataFrame, manager_name: str = "Manager") -> str:
    lines = [
        f"Hello {manager_name},",
        "",
        "This is an automated follow-up reminder from BRANDit Meeting Intelligence.",
        "",
        "The following tasks require follow-up:",
        "",
    ]

    for _, row in tasks_df.iterrows():
        bucket = compute_task_bucket(row.to_dict())
        lines.extend([
            f"Task ID: {row['id']}",
            f"Task: {row['task_description']}",
            f"Owner: {row.get('owner', 'TBD')}",
            f"Assignee Email: {row.get('owner_email', '')}",
            f"Priority: {row.get('priority', 'Medium')}",
            f"Meeting: {row.get('meeting_title', '')}",
            f"Due Date: {row.get('due_date', '')}",
            f"Status: {bucket}",
            f"Context: {row.get('context', '')}",
            "-" * 50,
        ])

    lines.extend([
        "",
        "Please review these tasks and follow up with the assigned team/person.",
        "",
        "Regards,",
        "BRANDit Meeting Intelligence",
    ])

    return "\n".join(lines)


def send_gmail_reminder(
    sender_gmail: str,
    gmail_app_password: str,
    receiver_email: str,
    tasks_df: pd.DataFrame,
    manager_name: str = "Manager",
) -> Tuple[bool, str]:
    """Send reminder email using Gmail SMTP + App Password."""
    sender_gmail = sender_gmail.strip()
    receiver_email = receiver_email.strip()

    if not sender_gmail or not gmail_app_password or not receiver_email:
        return False, "Sender Gmail, Gmail App Password and receiver email are required."
    if tasks_df.empty:
        return False, "No tasks require reminder right now."

    subject = f"Meeting Task Follow-up Reminder · {len(tasks_df)} task(s) need attention"
    body = build_reminder_body(tasks_df, manager_name=manager_name)

    msg = EmailMessage()
    msg["From"] = sender_gmail
    msg["To"] = receiver_email
    msg["Subject"] = subject
    msg.set_content(body)

    context = ssl.create_default_context()

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_gmail, gmail_app_password)
            server.send_message(msg)

        mark_reminders_sent(tasks_df["id"].astype(int).tolist())
        return True, f"Reminder email sent to {receiver_email} for {len(tasks_df)} task(s)."
    except smtplib.SMTPAuthenticationError:
        return False, "Gmail authentication failed. Use a Gmail App Password, not your normal Gmail password."
    except Exception as e:
        return False, f"Email sending failed: {e}"


def assignee_tasks_needing_reminders(manager_email: str = "") -> pd.DataFrame:
    """Pending tasks whose reminder date arrived and assignee email is available."""
    df = fetch_tasks(status_filter="All", manager_email=manager_email)
    if df.empty:
        return df

    today = today_date().isoformat()

    def should_remind(row) -> bool:
        if row["status"] == "Completed":
            return False
        if row["reminder_date"] > today:
            return False
        if pd.isna(row.get("owner_email")) or not str(row.get("owner_email", "")).strip():
            return False
        last = row.get("last_assignee_reminder_sent_at")
        if pd.isna(last) or not str(last).strip():
            return True
        try:
            return str(last)[:10] < today
        except Exception:
            return True

    return df[df.apply(should_remind, axis=1)].copy()


def mark_assignee_reminders_sent(task_ids: List[int]) -> None:
    if not task_ids:
        return

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    now = datetime.now().isoformat(timespec="seconds")
    cur.executemany(
        """
        UPDATE tasks
        SET last_assignee_reminder_sent_at = ?,
            assignee_reminder_count = COALESCE(assignee_reminder_count, 0) + 1
        WHERE id = ?
        """,
        [(now, int(task_id)) for task_id in task_ids],
    )
    conn.commit()
    conn.close()


def build_assignee_reminder_body(tasks_df: pd.DataFrame, assignee_name: str = "Team") -> str:
    lines = [
        f"Hello {assignee_name},",
        "",
        "This is an automated task follow-up reminder from BRANDit Meeting Intelligence.",
        "",
        "The following task(s) are assigned to you/your team and require action:",
        "",
    ]

    for _, row in tasks_df.iterrows():
        bucket = compute_task_bucket(row.to_dict())
        lines.extend([
            f"Task ID: {row['id']}",
            f"Task: {row['task_description']}",
            f"Priority: {row.get('priority', 'Medium')}",
            f"Meeting: {row.get('meeting_title', '')}",
            f"Assigned On: {row.get('assigned_on', '')}",
            f"Due Date: {row.get('due_date', '')}",
            f"Current Status: {bucket}",
            f"Context: {row.get('context', '')}",
            "-" * 50,
        ])

    lines.extend([
        "",
        "Please complete the task or update the manager if there is any blocker.",
        "",
        "Regards,",
        "BRANDit Meeting Intelligence",
    ])

    return "\n".join(lines)


def send_assignee_gmail_reminders(
    sender_gmail: str,
    gmail_app_password: str,
    tasks_df: pd.DataFrame,
    manager_email: str = "",
) -> Tuple[bool, str]:
    """Send separate reminder emails to each assigned person/team based on owner_email."""
    sender_gmail = sender_gmail.strip()
    manager_email = manager_email.strip()

    if not sender_gmail or not gmail_app_password:
        return False, "Sender Gmail and Gmail App Password are required."
    if tasks_df.empty:
        return False, "No assignee reminders are due right now."
    if "owner_email" not in tasks_df.columns:
        return False, "No owner_email column found. Add assignee emails first."

    tasks_df = tasks_df.copy()
    tasks_df["owner_email"] = tasks_df["owner_email"].fillna("").astype(str).str.strip()
    tasks_df = tasks_df[tasks_df["owner_email"].str.contains("@", regex=False)]
    if tasks_df.empty:
        return False, "No valid assignee email found for due tasks."

    context = ssl.create_default_context()
    sent_task_ids: List[int] = []
    sent_emails: List[str] = []

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_gmail, gmail_app_password)

            for assignee_email, group in tasks_df.groupby("owner_email"):
                assignee_name = str(group.iloc[0].get("owner", "Team")) or "Team"
                msg = EmailMessage()
                msg["From"] = sender_gmail
                msg["To"] = assignee_email
                if manager_email:
                    msg["Cc"] = manager_email
                msg["Subject"] = f"Task Follow-up Reminder · {len(group)} assigned task(s)"
                msg.set_content(build_assignee_reminder_body(group, assignee_name=assignee_name))
                server.send_message(msg)
                sent_task_ids.extend(group["id"].astype(int).tolist())
                sent_emails.append(assignee_email)

        mark_assignee_reminders_sent(sent_task_ids)
        unique_emails = sorted(set(sent_emails))
        return True, f"Assignee reminders sent to {len(unique_emails)} recipient(s): {', '.join(unique_emails)}."
    except smtplib.SMTPAuthenticationError:
        return False, "Gmail authentication failed. Use a Gmail App Password, not your normal Gmail password."
    except Exception as e:
        return False, f"Assignee email sending failed: {e}"


def send_webhook_notification(webhook_url: str, tasks_df: pd.DataFrame, channel_name: str = "Webhook") -> Tuple[bool, str]:
    """
    Works for simple Slack incoming webhook.
    For Microsoft Teams, use an Incoming Webhook or Power Automate webhook that accepts JSON text.
    """
    webhook_url = webhook_url.strip()
    if not webhook_url:
        return False, f"{channel_name} webhook URL is missing."
    if tasks_df.empty:
        return False, "No tasks require reminder right now."

    body = build_reminder_body(tasks_df, manager_name="Manager")
    payload = {"text": body}

    try:
        response = requests.post(webhook_url, json=payload, timeout=30)
        if response.status_code in (200, 201, 202, 204):
            return True, f"{channel_name} notification sent."
        return False, f"{channel_name} webhook failed ({response.status_code}): {response.text[:500]}"
    except Exception as e:
        return False, f"{channel_name} webhook error: {e}"


# ── Groq Whisper API Transcription ───────────────────────────────────────────
GROQ_TRANSCRIPTION_URL = "https://api.groq.com/openai/v1/audio/transcriptions"


def _raise_groq_audio_error(response: requests.Response) -> None:
    if response.ok:
        return
    try:
        detail = response.json()
    except Exception:
        detail = response.text
    raise RuntimeError(f"Groq Whisper API failed ({response.status_code}): {detail}")


def _to_plain_dict(value):
    if isinstance(value, dict):
        return value
    if hasattr(value, "model_dump"):
        return value.model_dump()
    if hasattr(value, "dict"):
        return value.dict()
    return value


def transcribe_with_groq_api(
    audio_path: str,
    groq_key: str,
    model_name: str = "whisper-large-v3-turbo",
    language: str = "en",
    prompt: str = "Meeting discussion with multiple speakers. Use clear punctuation.",
) -> dict:
    if not groq_key:
        raise RuntimeError("GROQ_API_KEY is missing.")

    headers = {"Authorization": f"Bearer {groq_key}"}

    data = {
        "model": model_name,
        "response_format": "verbose_json",
        "temperature": "0",
        "timestamp_granularities[]": "segment",
    }
    if prompt:
        data["prompt"] = prompt[:900]
    if language != "auto":
        data["language"] = language

    with open(audio_path, "rb") as audio_file:
        files = {"file": (Path(audio_path).name, audio_file, "audio/flac")}
        response = requests.post(
            GROQ_TRANSCRIPTION_URL,
            headers=headers,
            data=data,
            files=files,
            timeout=900,
        )

    _raise_groq_audio_error(response)
    result = _to_plain_dict(response.json())

    raw_segments = result.get("segments") or []
    segments = []
    full_parts = []

    for seg in raw_segments:
        seg = _to_plain_dict(seg)
        if not isinstance(seg, dict):
            continue
        text = (seg.get("text") or "").strip()
        if not text:
            continue
        start = float(seg.get("start", 0) or 0)
        end = float(seg.get("end", start) or start)
        if end <= start:
            end = start + 0.01
        segments.append({"start": start, "end": end, "text": text})
        full_parts.append(text)

    text = (result.get("text") or "").strip()
    if not segments and text:
        segments.append({"start": 0.0, "end": float(result.get("duration", 0) or 0), "text": text})
        full_parts.append(text)

    duration = float(result.get("duration", 0) or 0)
    if not duration and segments:
        duration = float(segments[-1]["end"])

    return {
        "full_text": " ".join(full_parts).strip() or text,
        "segments": segments,
        "words": [],
        "duration": duration,
        "language": result.get("language", language),
        "lang_prob": 0,
    }


# ── pyannoteAI API Diarization ───────────────────────────────────────────────
PYANNOTE_BASE_URL = "https://api.pyannote.ai/v1"


def _pyannote_headers(api_key: str, json_content: bool = True) -> dict:
    headers = {"Authorization": f"Bearer {api_key}"}
    if json_content:
        headers["Content-Type"] = "application/json"
    return headers


def _raise_pyannote_error(response: requests.Response, action: str) -> None:
    if response.ok:
        return
    try:
        detail = response.json()
    except Exception:
        detail = response.text
    raise RuntimeError(f"pyannoteAI {action} failed ({response.status_code}): {detail}")


def upload_audio_to_pyannote_media(audio_path: str, api_key: str) -> str:
    extension = Path(audio_path).suffix.lower().lstrip(".") or "flac"
    object_key = f"brandit-meetings/{datetime.now().strftime('%Y%m%d')}/{uuid.uuid4().hex}.{extension}"
    media_url = f"media://{object_key}"

    create_response = requests.post(
        f"{PYANNOTE_BASE_URL}/media/input",
        headers=_pyannote_headers(api_key),
        json={"url": media_url},
        timeout=60,
    )
    _raise_pyannote_error(create_response, "media URL creation")

    upload_url = create_response.json().get("url")
    if not upload_url:
        raise RuntimeError("pyannoteAI did not return a pre-signed upload URL.")

    with open(audio_path, "rb") as audio_file:
        upload_response = requests.put(
            upload_url,
            data=audio_file,
            headers={"Content-Type": "application/octet-stream"},
            timeout=600,
        )
    _raise_pyannote_error(upload_response, "audio upload")

    return media_url


def submit_pyannote_diarization_job(
    media_url: str,
    api_key: str,
    exact_speakers: Optional[int] = None,
    min_speakers: int = 1,
    max_speakers: int = 6,
    model: str = "precision-2",
) -> str:
    payload = {
        "url": media_url,
        "model": model,
        "exclusive": True,
        "turnLevelConfidence": False,
        "confidence": False,
        "transcription": False,
    }

    if exact_speakers:
        payload["numSpeakers"] = int(exact_speakers)
    else:
        payload["minSpeakers"] = int(min_speakers)
        payload["maxSpeakers"] = int(max_speakers)

    response = requests.post(
        f"{PYANNOTE_BASE_URL}/diarize",
        headers=_pyannote_headers(api_key),
        json=payload,
        timeout=60,
    )
    _raise_pyannote_error(response, "diarization job creation")

    data = response.json()
    job_id = data.get("jobId")
    if not job_id:
        raise RuntimeError(f"pyannoteAI did not return jobId: {data}")
    return job_id


def poll_pyannote_job(
    job_id: str,
    api_key: str,
    timeout_seconds: int = 1800,
    sleep_seconds: int = 8,
    progress_callback=None,
) -> dict:
    start_time = time.time()

    while True:
        response = requests.get(
            f"{PYANNOTE_BASE_URL}/jobs/{job_id}",
            headers=_pyannote_headers(api_key, json_content=False),
            timeout=60,
        )
        _raise_pyannote_error(response, "job polling")

        data = response.json()
        status = data.get("status", "unknown")

        if progress_callback:
            progress_callback(status)

        if status == "succeeded":
            return data
        if status in {"failed", "canceled"}:
            raise RuntimeError(f"pyannoteAI diarization job {status}: {data}")

        if time.time() - start_time > timeout_seconds:
            raise RuntimeError("pyannoteAI diarization timed out. Try a shorter audio file or rerun.")

        time.sleep(sleep_seconds)


def _read_time_value(value) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def parse_pyannote_turns(job_payload: dict) -> List[dict]:
    output = job_payload.get("output", {}) or {}
    diarization = (
        output.get("exclusiveDiarization")
        or output.get("exclusive_diarization")
        or output.get("diarization")
        or []
    )

    turns: List[dict] = []
    for item in diarization:
        if not isinstance(item, dict):
            continue

        segment = item.get("segment") if isinstance(item.get("segment"), dict) else {}
        start = item.get("start", segment.get("start"))
        end = item.get("end", segment.get("end"))
        speaker = (
            item.get("speaker")
            or item.get("label")
            or item.get("speakerLabel")
            or item.get("speaker_label")
            or "Unknown"
        )

        turns.append({
            "start": _read_time_value(start),
            "end": _read_time_value(end),
            "speaker_raw": str(speaker),
        })

    turns = [t for t in turns if t["end"] > t["start"]]
    turns.sort(key=lambda x: (x["start"], x["end"]))
    return turns


def diarize_audio_with_pyannote_api(
    audio_path: str,
    api_key: str,
    exact_speakers: Optional[int] = None,
    min_speakers: int = 1,
    max_speakers: int = 6,
    model: str = "precision-2",
    progress_callback=None,
) -> Tuple[List[dict], dict]:
    media_url = upload_audio_to_pyannote_media(audio_path, api_key)
    job_id = submit_pyannote_diarization_job(
        media_url=media_url,
        api_key=api_key,
        exact_speakers=exact_speakers,
        min_speakers=min_speakers,
        max_speakers=max_speakers,
        model=model,
    )
    job_payload = poll_pyannote_job(
        job_id=job_id,
        api_key=api_key,
        progress_callback=progress_callback,
    )
    turns = parse_pyannote_turns(job_payload)
    if not turns:
        raise RuntimeError("pyannoteAI completed, but no speaker turns were returned.")
    return turns, job_payload


def best_speaker_for_segment(start: float, end: float, turns: List[dict]) -> str:
    best = "Unknown"
    best_overlap = 0.0

    for t in turns:
        overlap_start = max(start, t["start"])
        overlap_end = min(end, t["end"])
        overlap = max(0.0, overlap_end - overlap_start)
        if overlap > best_overlap:
            best_overlap = overlap
            best = t["speaker_raw"]

    return best


def attach_speakers_to_segments(td: dict, turns: List[dict]) -> dict:
    for seg in td.get("segments", []):
        seg["speaker_raw"] = best_speaker_for_segment(seg["start"], seg["end"], turns)

    td["speaker_turns"] = turns
    return normalize_speaker_labels(td)


def normalize_speaker_labels(td: dict) -> dict:
    mapping: Dict[str, str] = {}

    for seg in td.get("segments", []):
        raw = seg.get("speaker_raw", "Unknown") or "Unknown"
        if raw == "Unknown":
            seg["speaker"] = "Unknown"
            continue
        if raw not in mapping:
            mapping[raw] = f"Speaker {len(mapping) + 1:02d}"
        seg["speaker"] = mapping[raw]

    for turn in td.get("speaker_turns", []):
        raw = turn.get("speaker_raw", "Unknown")
        turn["speaker"] = mapping.get(raw, "Unknown")

    td["speaker_map_raw_to_clean"] = mapping
    td["full_text"] = build_speaker_transcript(td, use_name=False)
    return td


def build_speaker_transcript(td: dict, speaker_names: Optional[Dict[str, str]] = None, use_name: bool = True) -> str:
    lines = []
    speaker_names = speaker_names or {}

    for seg in td.get("segments", []):
        speaker = seg.get("speaker", "Unknown")
        display_name = speaker_names.get(speaker, speaker) if use_name else speaker
        lines.append(f"[{fmt_time(seg.get('start', 0))}] {display_name}: {seg.get('text', '')}")

    return "\n".join(lines)


# ── Groq Meeting Analysis ────────────────────────────────────────────────────
ANALYSIS_SCHEMA = """\
{
  "meeting_title": "<inferred title>",
  "date": "<dd MMM YYYY>",
  "duration_estimate": "<e.g. 45 minutes>",
  "attendees_estimate": "<count or range>",
  "executive_summary": "<3-5 sentences>",
  "key_topics": [{"topic": "...", "summary": "..."}],
  "decisions_made": ["..."],
  "action_items": [
    {
      "id": 1,
      "action": "<clear task>",
      "owner": "<speaker/person/team or TBD>",
      "owner_email": "<email if clearly mentioned, otherwise blank>",
      "priority": "<High|Medium|Low>",
      "due_days": <integer>,
      "context": "<brief note explaining where the task came from>"
    }
  ],
  "risks_flagged": ["..."],
  "follow_up_questions": ["..."],
  "sentiment": "<Positive|Neutral|Mixed|Tense>",
  "meeting_effectiveness": "<N/10 — one sentence rationale>"
}"""

ESCAPED_SCHEMA = ANALYSIS_SCHEMA.replace("{", "{{").replace("}", "}}")

SYSTEM_PROMPT = """You are an expert meeting analyst and chief-of-staff AI.
Analyze the speaker-labelled transcript and reply ONLY with valid JSON matching this schema exactly.
No markdown fences, no preamble, no trailing text — pure JSON only.

Schema:
{schema}

Today: {{today}}

Critical task extraction rules:
- Extract every action item, commitment, task, deliverable and follow-up.
- Identify task_description/action, owner/team/person, owner_email if mentioned, due_days and priority.
- If the transcript says "in 5 days", due_days must be 5.
- If the transcript says "in 7 days", due_days must be 7.
- If the transcript says "by tomorrow", due_days must be 1.
- If the transcript says "next week", due_days must be 7.
- If no deadline is mentioned, use due_days = 5 and mark context as "No explicit deadline; default 5 days used."
- Use speaker labels/names when assigning action owners.
- If owner is unclear, use "TBD". If assignee email is not explicitly mentioned, keep owner_email blank.
- due_days must always be an integer.
- Flag blockers, risks, open questions, and follow-ups.
- Be precise and professional.""".format(schema=ESCAPED_SCHEMA)

GROQ_CHAT_COMPLETIONS_URL = "https://api.groq.com/openai/v1/chat/completions"


def _raise_groq_chat_error(response: requests.Response) -> None:
    if response.ok:
        return
    try:
        detail = response.json()
    except Exception:
        detail = response.text
    raise RuntimeError(f"Groq chat completion failed ({response.status_code}): {detail}")


def analyse_with_groq(transcript: str, filename: str, groq_key: str, model_name: str) -> dict:
    if not groq_key:
        raise RuntimeError("GROQ_API_KEY is missing.")

    today_text = datetime.now().strftime("%d %B %Y")
    system_prompt = SYSTEM_PROMPT.replace("{{today}}", today_text).replace("{today}", today_text)
    user_prompt = f"Filename: {filename}\n\nSpeaker-labelled transcript:\n{transcript[:18000]}"

    payload = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.1,
        "max_completion_tokens": 4096,
        "response_format": {"type": "json_object"},
    }

    response = requests.post(
        GROQ_CHAT_COMPLETIONS_URL,
        headers={
            "Authorization": f"Bearer {groq_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=900,
    )
    _raise_groq_chat_error(response)

    data = response.json()
    raw = (
        data.get("choices", [{}])[0]
        .get("message", {})
        .get("content", "")
        .strip()
    )

    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s*```$", "", raw)

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw, flags=re.DOTALL)
        if match:
            return json.loads(match.group(0))
        raise


# ── Excel Export ─────────────────────────────────────────────────────────────
def build_excel(analysis: dict, td: dict, filename: str, speaker_names: Optional[Dict[str, str]] = None) -> bytes:
    speaker_names = speaker_names or {}

    wb = Workbook()
    P = {
        "bg": "070712", "slate": "1A1A30", "sub": "111127", "text": "E0E0FF",
        "muted": "707090", "indigo": "6366F1", "high": "F43F5E", "med": "FBBF24",
        "low": "34D399", "row_a": "0A0A18", "row_b": "0D0D20", "white": "FFFFFF",
    }

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(size=10, bold=False, color=None): return Font(name="Arial", size=size, bold=bold, color=color or P["text"])
    def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="222244"),
        right=Side(style="thin", color="222244"),
        top=Side(style="thin", color="222244"),
        bottom=Side(style="thin", color="222244"),
    )

    def banner(ws, title, end_col="H"):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(f"A1:{end_col}2")
        c = ws["A1"]
        c.value = title
        c.fill = fill(P["indigo"])
        c.font = font(14, True, P["white"])
        c.alignment = center()
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 8

    def write_headers(ws, headers, widths, row=3):
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row, i, h)
            c.fill = fill(P["slate"])
            c.font = font(10, True, P["white"])
            c.alignment = center()
            c.border = border
            ws.column_dimensions[get_column_letter(i)].width = w

    # Summary
    ws = wb.active
    ws.title = "Summary"
    banner(ws, "BRANDit · MEETING INTELLIGENCE REPORT")
    rows = [
        ("Meeting Title", analysis.get("meeting_title", "Meeting")),
        ("Date", analysis.get("date", "")),
        ("Duration", analysis.get("duration_estimate", "")),
        ("Attendees", analysis.get("attendees_estimate", "")),
        ("Source File", filename),
        ("Executive Summary", analysis.get("executive_summary", "")),
    ]
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    for idx, (k, v) in enumerate(rows, 4):
        ws.cell(idx, 1, k).fill = fill(P["sub"])
        ws.cell(idx, 1).font = font(bold=True, color="93C5FD")
        ws.cell(idx, 1).alignment = left()
        ws.cell(idx, 2, v).fill = fill(P["bg"])
        ws.cell(idx, 2).font = font()
        ws.cell(idx, 2).alignment = left()
        ws.cell(idx, 1).border = border
        ws.cell(idx, 2).border = border

    # Action Tracker from current analysis
    wa = wb.create_sheet("Action Tracker")
    banner(wa, "ACTION POINTS TRACKER", "K")
    headers = ["#", "Action", "Owner", "Priority", "Assigned On", "Due Days", "Due Date", "Reminder Date", "Status", "Context", "Remarks"]
    widths = [6, 46, 22, 14, 16, 12, 16, 16, 14, 42, 28]
    write_headers(wa, headers, widths)

    prio_color = {"High": P["high"], "Medium": P["med"], "Low": P["low"]}
    for idx, a in enumerate(analysis.get("action_items", []), 1):
        row = 3 + idx
        due_days = safe_int(a.get("due_days", 5), default=5)
        assigned = today_date()
        due = assigned + timedelta(days=due_days)
        reminder = due - timedelta(days=1)
        if reminder < assigned:
            reminder = assigned

        vals = [
            idx,
            a.get("action", ""),
            a.get("owner", "TBD"),
            a.get("priority", "Medium"),
            assigned.strftime("%d %b %Y"),
            due_days,
            due.strftime("%d %b %Y"),
            reminder.strftime("%d %b %Y"),
            "Pending",
            a.get("context", ""),
            "",
        ]
        bg = P["row_a"] if idx % 2 else P["row_b"]
        for col, val in enumerate(vals, 1):
            c = wa.cell(row, col, val)
            c.fill = fill(bg)
            c.border = border
            c.alignment = center() if col in (1, 4, 5, 6, 7, 8, 9) else left()
            c.font = font(color=prio_color.get(str(val), P["text"]) if col == 4 else P["text"])

    # Transcript
    wt = wb.create_sheet("Transcript")
    banner(wt, "SPEAKER-LABELLED TRANSCRIPT", "D")
    write_headers(wt, ["Timestamp", "Speaker", "Text", "System Label"], [16, 22, 110, 24])

    for idx, seg in enumerate(td.get("segments", []), 1):
        row = 3 + idx
        speaker = seg.get("speaker", "Unknown")
        display = speaker_names.get(speaker, speaker)
        vals = [fmt_time(seg.get("start", 0)), display, seg.get("text", ""), seg.get("speaker_raw", "")]
        bg = P["row_a"] if idx % 2 else P["row_b"]
        for col, val in enumerate(vals, 1):
            c = wt.cell(row, col, val)
            c.fill = fill(bg)
            c.border = border
            c.font = font()
            c.alignment = left() if col == 3 else center()

    # Risks & Followups
    wr = wb.create_sheet("Risks & Follow-ups")
    banner(wr, "RISKS, BLOCKERS & OPEN QUESTIONS", "D")
    wr.column_dimensions["A"].width = 120
    row = 4
    for risk in analysis.get("risks_flagged", []):
        wr.cell(row, 1, f"Risk: {risk}")
        wr.cell(row, 1).fill = fill(P["row_a"])
        wr.cell(row, 1).font = font(color="FCA5A5")
        wr.cell(row, 1).alignment = left()
        wr.cell(row, 1).border = border
        row += 1
    for q in analysis.get("follow_up_questions", []):
        wr.cell(row, 1, f"Follow-up: {q}")
        wr.cell(row, 1).fill = fill(P["row_b"])
        wr.cell(row, 1).font = font(color="FDE68A")
        wr.cell(row, 1).alignment = left()
        wr.cell(row, 1).border = border
        row += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def build_tasks_excel(tasks_df: pd.DataFrame) -> bytes:
    output = BytesIO()
    export_df = tasks_df.copy()
    if not export_df.empty:
        cols = [
            "id", "meeting_title", "task_description", "owner", "owner_email", "priority",
            "assigned_on", "due_days", "due_date", "reminder_date",
            "status", "Bucket", "Days Left", "manager_email", "reminder_count", "assignee_reminder_count"
        ]
        export_df = export_df[[c for c in cols if c in export_df.columns]]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Task Dashboard", index=False)
    return output.getvalue()


# ── Session State ────────────────────────────────────────────────────────────
init_db()
for key, default in {
    "history": [],
    "current_session": None,
    "last_auto_check": "",
}.items():
    if key not in st.session_state:
        st.session_state[key] = default


# ── Secrets / Environment ────────────────────────────────────────────────────
GROQ_API_KEY = get_secret("GROQ_API_KEY")
PYANNOTE_API_KEY = (
    get_secret("PYANNOTE_API_KEY")
    or get_secret("PYANNOTEAI_API_KEY")
    or get_secret("PYANNOTE_API_TOKEN")
)
DEFAULT_GROQ_MODEL = get_secret("GROQ_MODEL", "llama-3.3-70b-versatile")
DEFAULT_WHISPER_API_MODEL = get_secret("WHISPER_API_MODEL", "whisper-large-v3-turbo")

VALID_GROQ_MODELS = ["llama-3.3-70b-versatile", "llama-3.1-8b-instant", "gemma2-9b-it"]
VALID_WHISPER_API_MODELS = ["whisper-large-v3-turbo", "whisper-large-v3"]
VALID_PYANNOTE_MODELS = ["precision-2", "community-1"]

groq_model_default = DEFAULT_GROQ_MODEL if DEFAULT_GROQ_MODEL in VALID_GROQ_MODELS else "llama-3.3-70b-versatile"
whisper_model_default = DEFAULT_WHISPER_API_MODEL if DEFAULT_WHISPER_API_MODEL in VALID_WHISPER_API_MODELS else "whisper-large-v3-turbo"


# ── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div class="brand-card">
        <h2>BRANDit Meeting Intelligence</h2>
        <p>Meeting minutes, action tracking, deadline calculation and automated follow-up reminders.</p>
        <div class="contact-strip">
            Prepared by <b>Kalpanasingh Chauhan</b><br>
            +91 8850159663<br>
            chauhankalpana2020@gmail.com
        </div>
    </div>
    """, unsafe_allow_html=True)

    page = st.radio(
        "Navigation",
        ["Process Meeting", "Task Dashboard", "Manual Task", "History"],
        label_visibility="collapsed",
    )

    st.divider()
    st.markdown("**API Setup**")
    st.markdown(f"Groq API: {'<span class=status-ok>Ready</span>' if GROQ_API_KEY else '<span class=status-bad>Missing</span>'}", unsafe_allow_html=True)
    st.markdown(f"pyannoteAI: {'<span class=status-ok>Ready</span>' if PYANNOTE_API_KEY else '<span class=status-bad>Missing / optional</span>'}", unsafe_allow_html=True)
    st.markdown(f"FFmpeg: {'<span class=status-ok>Ready</span>' if check_ffmpeg() else '<span class=status-bad>Missing</span>'}", unsafe_allow_html=True)

    st.divider()
    st.markdown("**Reminder Email Setup**")
    manager_name = st.text_input("Manager Name", value="Manager")
    manager_email = st.text_input("Manager Email / Receiver", value="")
    sender_gmail = st.text_input("Sender Gmail", value="", placeholder="yourgmail@gmail.com")
    gmail_app_password = st.text_input(
        "Gmail App Password",
        value="",
        type="password",
        help="Use Gmail App Password, not the normal Gmail login password.",
    )
    reminder_before_days = st.number_input("Reminder before due date", min_value=0, max_value=7, value=1, step=1)

    st.markdown("**Assigned Person Email Mapping**")
    assignee_mapping_text = st.text_area(
        "Owner/team to email",
        value="",
        placeholder="Team A=teama@company.com\nTeam B=teamb@company.com\nSpeaker 01=person@company.com",
        help="Use one mapping per line. The app uses this to send reminders to the person/team assigned the task.",
    )
    assignee_email_map = parse_assignee_email_mapping(assignee_mapping_text)
    send_assignee_reminders_enabled = st.checkbox("Also send reminders to assigned person/team", value=True)

    auto_email_check = st.checkbox(
        "Auto-check reminders when app opens/refreshes",
        value=False,
        help="This checks reminders when the Streamlit app is opened/refreshed. For true background reminders, run reminder_worker.py daily/hourly.",
    )

    st.divider()
    st.caption("Gmail reminders use smtp.gmail.com SSL port 465.")


# Auto-check reminders on app refresh/open
if auto_email_check and sender_gmail and gmail_app_password and manager_email:
    today_key = today_date().isoformat()
    if st.session_state.last_auto_check != today_key:
        due_df = tasks_needing_reminders(manager_email=manager_email)
        if not due_df.empty:
            ok, msg = send_gmail_reminder(sender_gmail, gmail_app_password, manager_email, due_df, manager_name)
            if ok:
                st.toast(msg)
            else:
                st.warning(msg)

        if send_assignee_reminders_enabled:
            assignee_due_df = assignee_tasks_needing_reminders(manager_email=manager_email)
            if not assignee_due_df.empty:
                ok, msg = send_assignee_gmail_reminders(
                    sender_gmail=sender_gmail,
                    gmail_app_password=gmail_app_password,
                    tasks_df=assignee_due_df,
                    manager_email=manager_email,
                )
                if ok:
                    st.toast(msg)
                else:
                    st.warning(msg)

        st.session_state.last_auto_check = today_key


# ── Page: Process Meeting ────────────────────────────────────────────────────
if page == "Process Meeting":
    st.markdown("""
    <div class="hero">
        <div class="eyebrow">BRANDit AI Prototype · Meeting Intelligence + Reminders</div>
        <h1>Meeting Intelligence & Follow-up Automation</h1>
        <p>
        Upload or record a meeting. The system creates transcript, minutes, action items,
        due dates, reminder dates, a task database and a manager follow-up dashboard.
        </p>
    </div>
    <div class="workflow">
        Workflow
        <span class="active">01 Upload / Record</span>
        <span>02 Transcribe</span>
        <span>03 Extract Tasks</span>
        <span>04 Save Tracker</span>
        <span>05 Send Reminder</span>
    </div>
    """, unsafe_allow_html=True)

    feature_cols = st.columns(3, gap="medium")
    with feature_cols[0]:
        st.markdown("""
        <div class="feature-card">
            <h3>Action Item Extraction</h3>
            <p>Detects task description, owner, priority and deadline from meeting transcript.</p>
        </div>
        """, unsafe_allow_html=True)
    with feature_cols[1]:
        st.markdown("""
        <div class="feature-card">
            <h3>Automatic Due Dates</h3>
            <p>If Team A has 5 days and Team B has 7 days, the due dates are calculated immediately.</p>
        </div>
        """, unsafe_allow_html=True)
    with feature_cols[2]:
        st.markdown("""
        <div class="feature-card">
            <h3>Reminder Dashboard</h3>
            <p>Shows pending, overdue, due today and upcoming follow-ups with email reminders.</p>
        </div>
        """, unsafe_allow_html=True)

    with st.expander("Processing Settings", expanded=False):
        c1, c2, c3 = st.columns(3)
        with c1:
            processing_mode = st.selectbox("Processing Mode", ["Fast API", "Balanced API", "Detailed API"], index=1)
            whisper_api_model = "whisper-large-v3" if processing_mode == "Detailed API" else whisper_model_default
            whisper_api_model = st.selectbox(
                "Transcription Model",
                VALID_WHISPER_API_MODELS,
                index=VALID_WHISPER_API_MODELS.index(whisper_api_model) if whisper_api_model in VALID_WHISPER_API_MODELS else 0,
            )
        with c2:
            groq_model = st.selectbox(
                "Groq Analysis Model",
                VALID_GROQ_MODELS,
                index=VALID_GROQ_MODELS.index(groq_model_default) if groq_model_default in VALID_GROQ_MODELS else 0,
            )
            language_label = st.selectbox("Meeting Language", ["English", "Auto detect"], index=0)
            language = "en" if language_label == "English" else "auto"
        with c3:
            enable_diarization = st.checkbox("Create speaker-wise transcript", value=True)
            pyannote_model = st.selectbox("Speaker diarization model", VALID_PYANNOTE_MODELS, index=0)
            known_count = st.checkbox("I know the number of speakers", value=True)
            if known_count:
                exact_speakers = st.number_input("Number of speakers", min_value=1, max_value=12, value=2, step=1)
                min_speakers = max_speakers = int(exact_speakers)
            else:
                exact_speakers = None
                min_speakers = st.number_input("Minimum speakers", min_value=1, max_value=12, value=1, step=1)
                max_speakers = st.number_input("Maximum speakers", min_value=int(min_speakers), max_value=12, value=max(2, int(min_speakers)), step=1)

    left, right = st.columns([1.08, 0.92], gap="large")

    with left:
        st.markdown('<div class="card"><div class="card-title">Meeting Input</div>', unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Upload meeting audio/video",
            type=["mp3", "mp4", "wav", "m4a", "ogg", "webm", "flac", "mkv", "avi", "mov"],
        )

        recorded_audio = None
        if hasattr(st, "audio_input"):
            recorded_audio = st.audio_input("Or record meeting audio")

        meeting_file = uploaded or recorded_audio

        if meeting_file:
            size_mb = getattr(meeting_file, "size", 0) / 1024 / 1024
            st.success(f"Ready for analysis: {meeting_file.name} ({size_mb:.1f} MB)")

        st.markdown('</div>', unsafe_allow_html=True)

        missing = []
        if not GROQ_API_KEY:
            missing.append("GROQ_API_KEY")
        if enable_diarization and not PYANNOTE_API_KEY:
            missing.append("PYANNOTE_API_KEY")
        if not check_ffmpeg():
            missing.append("FFmpeg")

        if not manager_email.strip():
            st.info("Add manager email in the sidebar so extracted tasks are stored with the reminder receiver.")

        run = st.button("Analyze Meeting & Create Tasks", disabled=(meeting_file is None or bool(missing)))
        if missing:
            st.warning("Setup required: " + ", ".join(missing))

    with right:
        st.markdown('<div class="card"><div class="card-title">How the manager gets automatic follow-up</div>', unsafe_allow_html=True)
        st.write(
            "When action items are extracted, each task is saved in SQLite with `assigned_on`, `due_days`, "
            "`due_date`, and `reminder_date`. The dashboard checks tasks where `reminder_date <= today` "
            "and sends the manager an email reminder. Overdue and due-today tasks are highlighted automatically."
        )
        st.code(
            "Team A task → due_days = 5 → due_date = today + 5 → reminder_date = due_date - 1\n"
            "Team B task → due_days = 7 → due_date = today + 7 → reminder_date = due_date - 1",
            language="text",
        )
        st.markdown('</div>', unsafe_allow_html=True)

    if meeting_file and run:
        suffix = Path(meeting_file.name).suffix or ".tmp"
        tmp_input = None
        tmp_audio = None

        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
                f.write(meeting_file.read())
                tmp_input = f.name

            with tempfile.NamedTemporaryFile(delete=False, suffix=".flac") as f:
                tmp_audio = f.name

            progress = st.progress(0, "Preparing the recording...")
            prepare_audio_for_apis(tmp_input, tmp_audio)

            progress.progress(15, "Creating transcript with Groq Whisper API...")
            td = transcribe_with_groq_api(tmp_audio, GROQ_API_KEY or "", whisper_api_model, language)

            pyannote_payload = None
            if enable_diarization:
                progress.progress(45, "Running speaker tracking...")

                def api_status(status: str):
                    progress.progress(55, f"Speaker tracking status: {status}...")

                turns, pyannote_payload = diarize_audio_with_pyannote_api(
                    audio_path=tmp_audio,
                    api_key=PYANNOTE_API_KEY or "",
                    exact_speakers=int(exact_speakers) if exact_speakers else None,
                    min_speakers=int(min_speakers),
                    max_speakers=int(max_speakers),
                    model=pyannote_model,
                    progress_callback=api_status,
                )
                td = attach_speakers_to_segments(td, turns)
            else:
                for seg in td.get("segments", []):
                    seg["speaker_raw"] = "Unknown"
                    seg["speaker"] = "Speaker 01"
                td["speaker_turns"] = []
                td["speaker_map_raw_to_clean"] = {}
                td["full_text"] = build_speaker_transcript(td, use_name=False)

            progress.progress(75, "Extracting minutes, action items, owners and timelines...")
            transcript_for_llm = build_speaker_transcript(td, use_name=False)
            analysis = analyse_with_groq(transcript_for_llm, meeting_file.name, GROQ_API_KEY or "", groq_model)

            progress.progress(85, "Saving tasks to database...")
            inserted = save_tasks_from_analysis(
                analysis=analysis,
                source_file=meeting_file.name,
                manager_email=manager_email,
                reminder_before_days=int(reminder_before_days),
                assignee_email_map=assignee_email_map,
            )

            progress.progress(92, "Preparing Excel report...")
            xlsx = build_excel(analysis, td, meeting_file.name)

            progress.progress(100, "Analysis complete.")

            session = {
                "id": len(st.session_state.history) + 1,
                "filename": meeting_file.name,
                "timestamp": datetime.now().strftime("%d %b %Y · %H:%M"),
                "td": td,
                "analysis": analysis,
                "xlsx": xlsx,
                "model": groq_model,
                "whisper_api_model": whisper_api_model,
                "pyannote_model": pyannote_model,
                "pyannote_payload": pyannote_payload,
                "tasks_created": inserted,
            }
            st.session_state.history.append(session)
            st.session_state.current_session = session
            st.success(f"Meeting analysis completed. {inserted} task(s) saved to reminder tracker.")

        except json.JSONDecodeError as e:
            st.error(f"The analysis response was incomplete: {e}. Please rerun the meeting analysis.")
        except Exception as e:
            st.error(f"Error: {e}")
        finally:
            for p in [tmp_input, tmp_audio]:
                if p and os.path.exists(p):
                    try:
                        os.remove(p)
                    except Exception:
                        pass

    # Results
    session = st.session_state.current_session
    if session:
        td = session["td"]
        analysis = session["analysis"]
        segments = td.get("segments", [])
        speaker_labels = sorted({s.get("speaker", "Unknown") for s in segments})

        st.markdown('<div class="metric-strip">', unsafe_allow_html=True)
        metrics = [
            (len(segments), "Transcript Lines"),
            (fmt_time(td.get("duration", 0)), "Duration"),
            (len(speaker_labels), "Speakers"),
            (len(analysis.get("action_items", [])), "Action Points"),
            (session.get("tasks_created", 0), "Tasks Stored"),
        ]
        for value, label in metrics:
            st.markdown(f'<div class="metric-box"><div class="metric-num">{value}</div><div class="metric-label">{label}</div></div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        speaker_names: Dict[str, str] = {}
        if speaker_labels:
            st.markdown('<div class="section-header">Speaker Names</div>', unsafe_allow_html=True)
            cols = st.columns(min(4, max(1, len(speaker_labels))))
            for i, sp in enumerate(speaker_labels):
                with cols[i % len(cols)]:
                    default = sp if sp != "Unknown" else "Unknown"
                    speaker_names[sp] = st.text_input(f"Name for {sp}", value=default, key=f"name_{session['id']}_{sp}")

        named_transcript = build_speaker_transcript(td, speaker_names=speaker_names, use_name=True)

        tabs = st.tabs(["Minutes", "Transcript", "Action Items", "Risks & Follow-ups", "Export"])

        with tabs[0]:
            st.markdown('<div class="card"><div class="card-title">Executive Summary</div>', unsafe_allow_html=True)
            st.write(analysis.get("executive_summary", ""))
            st.markdown('</div>', unsafe_allow_html=True)

            topics = analysis.get("key_topics", [])
            if topics:
                st.markdown('<div class="section-header">Key Topics</div>', unsafe_allow_html=True)
                for t in topics:
                    st.markdown(f"**{t.get('topic','')}**")
                    st.write(t.get("summary", ""))

            decisions = analysis.get("decisions_made", [])
            if decisions:
                st.markdown('<div class="section-header">Decisions Made</div>', unsafe_allow_html=True)
                for d in decisions:
                    st.success(d)

        with tabs[1]:
            st.markdown(f"<div class='small-muted'>Language: <b>{html.escape(str(td.get('language',''))).upper()}</b> · Duration: <b>{fmt_time(td.get('duration',0))}</b> · Speakers: <b>{len(speaker_labels)}</b></div>", unsafe_allow_html=True)
            rows = []
            for seg in segments[:300]:
                speaker = seg.get("speaker", "Unknown")
                display = speaker_names.get(speaker, speaker)
                rows.append(
                    '<div class="seg-row">'
                    f'<span class="seg-time">{html.escape(fmt_time(seg.get("start", 0)))}</span>'
                    f'<span class="seg-spk">{html.escape(display)}</span>'
                    f'<span class="seg-text">{html.escape(seg.get("text", ""))}</span>'
                    '</div>'
                )
            if len(segments) > 300:
                rows.append(f"<p class='small-muted' style='text-align:center;'>… {len(segments)-300} more segments in export</p>")
            st.markdown('<div class="transcript-box">' + ''.join(rows) + '</div>', unsafe_allow_html=True)
            st.text_area("Plain transcript", named_transcript, height=240, label_visibility="collapsed")

        with tabs[2]:
            actions = analysis.get("action_items", [])
            if actions:
                table = []
                for a in actions:
                    due_days = safe_int(a.get("due_days", 5), default=5)
                    assigned = today_date()
                    due = assigned + timedelta(days=due_days)
                    reminder = due - timedelta(days=int(reminder_before_days))
                    if reminder < assigned:
                        reminder = assigned
                    owner_for_email = a.get("owner", "TBD")
                    owner_email_for_display = (a.get("owner_email") or a.get("assignee_email") or lookup_owner_email(owner_for_email, assignee_email_map))
                    table.append({
                        "#": a.get("id", ""),
                        "Action": a.get("action", ""),
                        "Owner": owner_for_email,
                        "Assignee Email": owner_email_for_display,
                        "Priority": a.get("priority", "Medium"),
                        "Assigned On": assigned.strftime("%d %b %Y"),
                        "Due Days": due_days,
                        "Due Date": due.strftime("%d %b %Y"),
                        "Reminder Date": reminder.strftime("%d %b %Y"),
                        "Status": "Pending",
                        "Context": a.get("context", ""),
                    })
                st.dataframe(pd.DataFrame(table), use_container_width=True, hide_index=True)
            else:
                st.info("No action items detected.")

        with tabs[3]:
            risks = analysis.get("risks_flagged", [])
            followups = analysis.get("follow_up_questions", [])
            if risks:
                st.markdown('<div class="section-header">Risks</div>', unsafe_allow_html=True)
                for r in risks:
                    st.error(r)
            if followups:
                st.markdown('<div class="section-header">Follow-up Questions</div>', unsafe_allow_html=True)
                for q in followups:
                    st.warning(q)
            if not risks and not followups:
                st.info("No major risks or follow-up questions detected.")

        with tabs[4]:
            base_name = Path(session["filename"]).stem
            c1, c2, c3 = st.columns(3)

            with c1:
                xlsx_named = build_excel(analysis, td, session["filename"], speaker_names=speaker_names)
                st.download_button(
                    "Download Meeting Excel",
                    data=xlsx_named,
                    file_name=f"brandit_meeting_report_{base_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            with c2:
                minutes = (
                    f"MEETING MINUTES\n{'='*60}\n"
                    f"{analysis.get('meeting_title','Meeting')}\n"
                    f"Date: {analysis.get('date','')}\n"
                    f"Duration: {analysis.get('duration_estimate','')}\n\n"
                    f"EXECUTIVE SUMMARY\n{'-'*30}\n{analysis.get('executive_summary','')}\n\n"
                    f"DECISIONS\n{'-'*30}\n" + "\n".join(f"- {d}" for d in analysis.get("decisions_made", [])) + "\n\n"
                    f"ACTION ITEMS\n{'-'*30}\n" + "\n".join(f"[{a.get('priority','')}] {a.get('action','')} — {a.get('owner','TBD')} — due in {a.get('due_days',5)} days" for a in analysis.get("action_items", []))
                )
                st.download_button("Download Minutes TXT", data=minutes, file_name=f"brandit_minutes_{base_name}.txt", mime="text/plain")

            with c3:
                st.download_button("Download Transcript", data=named_transcript, file_name=f"brandit_transcript_{base_name}.txt", mime="text/plain")


# ── Page: Task Dashboard ─────────────────────────────────────────────────────
elif page == "Task Dashboard":
    st.markdown("""
    <div class="hero">
        <div class="eyebrow">BRANDit AI Prototype · Follow-up Control Center</div>
        <h1>Task Dashboard & Reminder System</h1>
        <p>
        This dashboard is the answer to the business question:
        the manager does not need to manually track 5-day or 7-day tasks.
        The app stores each task with its calculated due date and reminder date,
        then checks which reminders are due.
        </p>
    </div>
    """, unsafe_allow_html=True)

    all_df = fetch_tasks(status_filter="All", manager_email=manager_email if manager_email else "")
    if all_df.empty:
        st.info("No tasks found yet. Process a meeting or add a manual task.")
    else:
        pending_count = int((all_df["status"] != "Completed").sum())
        overdue_count = int((all_df["Bucket"] == "Overdue").sum())
        today_count = int((all_df["Bucket"] == "Due Today").sum())
        soon_count = int((all_df["Bucket"] == "Due Soon").sum())
        reminder_count = len(tasks_needing_reminders(manager_email=manager_email if manager_email else ""))
        assignee_reminder_count = len(assignee_tasks_needing_reminders(manager_email=manager_email if manager_email else ""))

        st.markdown('<div class="metric-strip">', unsafe_allow_html=True)
        for value, label in [
            (pending_count, "Pending"),
            (overdue_count, "Overdue"),
            (today_count, "Due Today"),
            (soon_count, "Due Soon"),
            (reminder_count, "Manager Reminder"),
            (assignee_reminder_count, "Assignee Reminder"),
        ]:
            st.markdown(f'<div class="metric-box"><div class="metric-num">{value}</div><div class="metric-label">{label}</div></div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            status_filter = st.selectbox("Filter", ["All", "Pending", "Overdue", "Due Today", "Due Soon", "Upcoming", "Completed"])
        with c2:
            show_email = st.text_input("Filter by manager email", value=manager_email)
        with c3:
            st.write("")
            st.write("")
            refresh = st.button("Refresh Dashboard")
            if refresh:
                st.rerun()

        df = fetch_tasks(status_filter="All" if status_filter == "Pending" else status_filter, manager_email=show_email)
        if status_filter == "Pending" and not df.empty:
            df = df[df["status"] != "Completed"]

        display_cols = [
            "id", "Bucket", "task_description", "owner", "owner_email", "priority", "meeting_title",
            "assigned_on", "due_days", "due_date", "reminder_date", "status",
            "Days Left", "reminder_count", "assignee_reminder_count"
        ]
        st.dataframe(
            df[[c for c in display_cols if c in df.columns]],
            use_container_width=True,
            hide_index=True,
        )

        st.markdown('<div class="section-header">Send Manager + Assignee Follow-up Reminders</div>', unsafe_allow_html=True)
        reminder_df = tasks_needing_reminders(manager_email=show_email)
        assignee_reminder_df = assignee_tasks_needing_reminders(manager_email=show_email)

        if reminder_df.empty and assignee_reminder_df.empty:
            st.success("No manager or assignee reminders are due right now.")
        else:
            if not reminder_df.empty:
                st.warning(f"{len(reminder_df)} task(s) need manager reminder.")
                st.dataframe(
                    reminder_df[[c for c in display_cols if c in reminder_df.columns]],
                    use_container_width=True,
                    hide_index=True,
                )

            if not assignee_reminder_df.empty:
                st.warning(f"{len(assignee_reminder_df)} task(s) need assignee reminder.")
                st.dataframe(
                    assignee_reminder_df[[c for c in display_cols if c in assignee_reminder_df.columns]],
                    use_container_width=True,
                    hide_index=True,
                )

            b1, b2, b3 = st.columns(3)
            with b1:
                if st.button("Send Manager Gmail Reminder"):
                    ok, msg = send_gmail_reminder(
                        sender_gmail=sender_gmail,
                        gmail_app_password=gmail_app_password,
                        receiver_email=show_email or manager_email,
                        tasks_df=reminder_df,
                        manager_name=manager_name,
                    )
                    st.success(msg) if ok else st.error(msg)

            with b2:
                if st.button("Send Assignee Gmail Reminders"):
                    ok, msg = send_assignee_gmail_reminders(
                        sender_gmail=sender_gmail,
                        gmail_app_password=gmail_app_password,
                        tasks_df=assignee_reminder_df,
                        manager_email=show_email or manager_email,
                    )
                    st.success(msg) if ok else st.error(msg)

            with b3:
                if st.button("Send Both Manager + Assignee"):
                    messages = []
                    ok1, msg1 = send_gmail_reminder(
                        sender_gmail=sender_gmail,
                        gmail_app_password=gmail_app_password,
                        receiver_email=show_email or manager_email,
                        tasks_df=reminder_df,
                        manager_name=manager_name,
                    )
                    messages.append(msg1)
                    ok2, msg2 = send_assignee_gmail_reminders(
                        sender_gmail=sender_gmail,
                        gmail_app_password=gmail_app_password,
                        tasks_df=assignee_reminder_df,
                        manager_email=show_email or manager_email,
                    )
                    messages.append(msg2)
                    if ok1 or ok2:
                        st.success("\n".join(messages))
                    else:
                        st.error("\n".join(messages))

            with st.expander("Optional Slack / Teams manager notification"):
                b4, b5 = st.columns(2)
                with b4:
                    slack_webhook = st.text_input("Slack webhook URL", type="password", key="slack_webhook")
                    if st.button("Send Slack Webhook"):
                        ok, msg = send_webhook_notification(slack_webhook, reminder_df, "Slack")
                        st.success(msg) if ok else st.error(msg)
                with b5:
                    teams_webhook = st.text_input("Teams webhook URL", type="password", key="teams_webhook")
                    if st.button("Send Teams Webhook"):
                        ok, msg = send_webhook_notification(teams_webhook, reminder_df, "Teams")
                        st.success(msg) if ok else st.error(msg)

        st.markdown('<div class="section-header">Update Task Status</div>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            task_id = st.number_input("Task ID", min_value=1, step=1)
        with c2:
            new_status = st.selectbox("New Status", ["Pending", "Completed"])
        with c3:
            st.write("")
            st.write("")
            if st.button("Update Status"):
                update_task_status(int(task_id), new_status)
                st.success(f"Task {task_id} updated to {new_status}.")
                st.rerun()

        c1, c2 = st.columns([1, 2])
        with c1:
            if st.button("Delete Selected Task"):
                delete_task(int(task_id))
                st.warning(f"Task {task_id} deleted.")
                st.rerun()

        st.markdown('<div class="section-header">Export Task Tracker</div>', unsafe_allow_html=True)
        st.download_button(
            "Download Task Dashboard Excel",
            data=build_tasks_excel(all_df),
            file_name=f"brandit_task_dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# ── Page: Manual Task ────────────────────────────────────────────────────────
elif page == "Manual Task":
    st.markdown("""
    <div class="hero">
        <div class="eyebrow">BRANDit AI Prototype · Manual Follow-up</div>
        <h1>Add Manual Task</h1>
        <p>Add a task manually when the manager wants to track something outside the meeting transcript.</p>
    </div>
    """, unsafe_allow_html=True)

    with st.form("manual_task_form"):
        task_description = st.text_area("Task Description", placeholder="Team A to complete market research")
        owner = st.text_input("Assigned Team / Person", placeholder="Team A")
        owner_email = st.text_input("Assigned Person/Team Email", placeholder="teama@company.com")
        meeting_title = st.text_input("Meeting / Project", value="Manual Follow-up")
        priority = st.selectbox("Priority", ["High", "Medium", "Low"], index=1)
        due = st.date_input("Due Date", value=today_date() + timedelta(days=5))
        reminder = st.date_input("Reminder Date", value=today_date() + timedelta(days=4))
        context = st.text_area("Context / Notes", placeholder="Optional")
        submitted = st.form_submit_button("Save Task")

        if submitted:
            if not task_description.strip():
                st.error("Task description is required.")
            else:
                add_manual_task(
                    meeting_title=meeting_title,
                    task_description=task_description,
                    owner=owner,
                    owner_email=owner_email,
                    priority=priority,
                    due_date_value=due,
                    reminder_date_value=reminder,
                    manager_email=manager_email,
                    context=context,
                )
                st.success("Manual task saved.")


# ── Page: History ────────────────────────────────────────────────────────────
else:
    st.markdown("""
    <div class="hero">
        <div class="eyebrow">BRANDit AI Prototype · Session Archive</div>
        <h1>Meeting History</h1>
        <p>Review meeting reports processed during this Streamlit session.</p>
    </div>
    """, unsafe_allow_html=True)

    if not st.session_state.history:
        st.info("No meetings processed in this session yet.")
    else:
        for sess in reversed(st.session_state.history):
            with st.expander(f"{sess['filename']} · {sess['timestamp']}"):
                st.write(sess["analysis"].get("executive_summary", ""))
                st.write(f"Tasks created: {sess.get('tasks_created', 0)}")
                st.download_button(
                    "Download Meeting Excel Report",
                    data=sess["xlsx"],
                    file_name=f"brandit_meeting_report_{Path(sess['filename']).stem}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"hist_xlsx_{sess['id']}",
                )

st.markdown("""
<div class="footer-bar">
    Prepared by <b>Kalpanasingh Chauhan</b> &nbsp; · &nbsp; +91 8850159663 &nbsp; · &nbsp; chauhankalpana2020@gmail.com
</div>
""", unsafe_allow_html=True)
